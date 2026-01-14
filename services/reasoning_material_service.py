import json
import os
import posixpath
import shutil
from PIL import Image as PILImage
from flask import current_app
from models import ReasoningMaterial, db
from utils import extract_first_frame
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict
from loguru import logger as log


class ReasoningMaterialService:
    IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
    VIDEO_EXTS = {".mp4", ".avi", ".mov", ".mkv", ".flv"}
    BASE_DIR = "/"

    def __init__(self):
        pass

    def save_material_file(
        self,
        src_path: str,
        project_id: int,
        upload_base: str,  # e.g. static/uploads/{project_id}
        material_type: int,  # 1=image, 2=video, 3=rtsp
        source_type: str = "",
        origin_filename: str = None,
        rtsp_url: str = None,
        start_time: str = None,
        end_time: str = None,
    ):
        """
        统一处理素材文件保存、信息提取、数据库记录创建
        :param src_path: 源文件路径（本地文件/RTSP地址）
        :param project_id: 项目ID
        :param upload_base: 上传基础目录（static/uploads/{project_id}）
        :param material_type: 素材类型 1=图片 2=视频 3=RTSP
        :param source_type: 来源类型（可选）
        :param origin_filename: 原始文件名（可选，优先使用）
        :param rtsp_url: RTSP地址（仅material_type=3时使用）
        :param start_time: RTSP开始时间（仅material_type=3时使用）
        :param end_time: RTSP结束时间（仅material_type=3时使用）
        :return: 创建的ReasoningMaterial对象
        """
        # 1. 处理不同素材类型的基础信息
        if material_type == 3:  # RTSP特殊处理
            filename = origin_filename or "RTSP_STREAM"
            final_filename = filename
            size_bytes = 0
            width = height = None
            relative_path = rtsp_url or src_path

            # 提取RTSP第一帧作为封面
            cover_abs_path = extract_first_frame(src_path, upload_base)
            try:
                img = PILImage.open(cover_abs_path)
                width, height = img.size
            except Exception as e:
                current_app.logger.warning(
                    f"[WARN] 读取RTSP封面失败: {cover_abs_path}, {e}"
                )
        else:  # 图片/视频处理
            # 生成不重复的文件名
            filename = origin_filename or os.path.basename(src_path)
            name, ext = os.path.splitext(filename)
            final_filename = f"{name}_{int(time.time() * 1000)}{ext}"

            # 保存文件到目标路径
            target_path = os.path.join(upload_base, final_filename)
            os.makedirs(upload_base, exist_ok=True)
            shutil.copy2(src_path, target_path)

            # 获取文件基础信息
            size_bytes = os.path.getsize(target_path)
            width = height = None

            # 提取图片/视频尺寸
            if material_type == 1:  # 图片
                try:
                    img = PILImage.open(target_path)
                    width, height = img.size
                except Exception as e:
                    current_app.logger.warning(
                        f"[WARN] 读取图片失败: {target_path}, {e}"
                    )
            elif material_type == 2:  # 视频
                # 提取视频第一帧作为封面并获取尺寸
                cover_abs_path = extract_first_frame(target_path, upload_base)
                try:
                    img = PILImage.open(cover_abs_path)
                    width, height = img.size
                except Exception as e:
                    current_app.logger.warning(
                        f"[WARN] 读取视频封面失败: {cover_abs_path}, {e}"
                    )

            # 生成相对路径（URL用）
            relative_path = os.path.relpath(target_path, "static")
            relative_path = posixpath.join(*relative_path.split(os.sep))

        if material_type != 3:  # 非RTSP需要处理封面路径
            if material_type == 1:  # 图片用自身作为封面
                cover_abs_path = os.path.join("static", relative_path)
            # 视频封面已在上面提取，RTSP封面也已提取

        # 转换封面路径为相对路径（数据库存储）
        cover_relative_path = os.path.relpath(cover_abs_path, "static")
        cover_relative_path = posixpath.join(*cover_relative_path.split(os.sep))

        material = ReasoningMaterial(
            filename=final_filename,
            path=relative_path,
            project_id=project_id,
            status=0,
            source_type=source_type,
            material_type=material_type,
            width=width,
            height=height,
            size_bytes=size_bytes,
            cover_image=cover_relative_path,
            start_time=start_time or None,
            end_time=end_time or None,
        )

        db.session.add(material)

        # if material_type in [2, 3] and os.path.exists(cover_abs_path):
        #     try:
        #         os.remove(cover_abs_path)
        #     except Exception as e:
        #         current_app.logger.warning(
        #             f"[WARN] 清理临时封面失败: {cover_abs_path}, {e}"
        #         )

        return material

    @staticmethod
    def build_excel(data, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "统计结果"

        # 样式设置
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 表头
        headers = ["项目名称", "分类名称"] + data["labels"] + ["合计"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)  # 直接写入文字
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 数据行
        for row_idx, row in enumerate(data["rows"], start=2):
            ws.cell(row=row_idx, column=1, value=data["project_name"])
            ws.cell(row=row_idx, column=2, value=row["category"])
            for col_idx, label in enumerate(data["labels"], start=3):
                ws.cell(row=row_idx, column=col_idx, value=row["labels"].get(label, 0))
            ws.cell(row=row_idx, column=3 + len(data["labels"]), value=row["total"])

            # 给每个单元格加边框和居中
            for col_num in range(1, 3 + len(data["labels"]) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.alignment = center_align
                cell.border = thin_border

        # 自适应列宽
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 4

        wb.save(file_path)

    @staticmethod
    def aggregate_by_category(materials, project_name):
        category_map = defaultdict(lambda: defaultdict(int))
        all_labels = set()

        for m in materials:
            if not m.result:
                continue

            result_json = m.result
            success = result_json.get("success", False)
            if not success:
                continue
            class_counts = result_json.get("result", {}).get("class_counts", {})
            if not class_counts:
                continue

            if m.material_type == 1:
                category = m.source_type or m.filename  # 如果图片没有分类就使用名称
            elif m.material_type == 2:
                category = os.path.join(m.source_type or "", m.filename)  # 视频
            else:
                category = m.filename  # RTSP 用 filename
            for label, cnt in class_counts.items():
                category_map[category][label] += cnt
                all_labels.add(label)

        rows = []
        for category, label_map in category_map.items():
            total = sum(label_map.values())
            rows.append({"category": category, "labels": label_map, "total": total})

        return {
            "project_name": project_name,
            "labels": sorted(all_labels),
            "rows": rows,
        }

    def update_material(self, material_id, result=None, status=None):
        """
        素材推理过程实时更新结果可以使用这个方法

        :param self: 说明
        :param material_id: 说明
        :param result: 说明
        :param status: 说明
        """
        material = ReasoningMaterial.query.get(material_id)
        if not material:
            log.warning(f"素材ID {material_id} 不存在，跳过更新")
            return

        if result is not None:
            original_result = material.result.copy()
            original_result["result"] = result
            material.result = original_result

        if status is not None:
            material.status = status
        material.updated_at = time.strftime("%Y-%m-%d %H:%M:%S")

        db.session.commit()
        log.debug(f"素材 {material_id} 数据更新成功：result={result}")
        return


reasoning_material_service = ReasoningMaterialService()
